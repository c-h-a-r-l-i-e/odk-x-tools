#!\\usr\\bin\\python3
import sys
from openpyxl import load_workbook
import os

# The product of these two is the number of characters we can log, errors appear for more than 
# 60,000 characters per log column
LOG_COLUMN_NUMBER = 20
CHARS_PER_LOG = 60000

#List tables we mark as linked to persons - i.e. the forms for these tables will be included in the list of buttons on a person's page
PERSON_TABLES = ['survey', 'sample']

class Form:
    def __init__(self, filename):
        self.filename = filename
        self.workbook = load_workbook(filename = filename)
        self.formId = self.getFormId()
        self.tableId = self.getTableId()
        self.title = self.getTitle()


    def convert(self):
        try:
            ws = self.workbook['survey'] #Catch error when survey does not exist
        except:
            raise ValueError("Workbook {} must have a sheet named 'survey'".format(self.filename))
            
        typeIndex = -1
        nameIndex = -1
        calculationIndex = -1
        clauseIndex = -1

        for i in range(len(ws[1])): #Scan each cell in the top row to find the columns containing type, name and calculation
            if ws[1][i].value == "type":
                typeIndex = i + 1

            elif ws[1][i].value == "name":
                nameIndex = i + 1
        
            elif ws[1][i].value == "calculation":
                calculationIndex = i + 1

            elif ws[1][i].value == "clause":
                clauseIndex = i + 1

        if typeIndex < 0 or nameIndex <= 0:
            raise ValueError("Workbook provided must have both 'type' and 'name' columns.")

        if calculationIndex < 0: #Insert a calculation column
            ws.cell(1, len(ws[1]) + 1, "calculation")
            calculationIndex = len(ws[i])

        clauseExists = clauseIndex != -1 

        
        # Iterate through fields, adding logging as we go
        offset = 0 #value to allow us to skip over any new rows we add
        #screenCalc used to store a list of calculations that need to be added after end screen - we don't want multiple assigns on the same screen
        screenCalcs = []
        inScreen = False
        for i in range(2, ws.max_row + 1):
            i += offset
            row = ws[i]
            

            if inScreen:
                if clauseExists and ws.cell(i, clauseIndex).value != None and 'end screen' in ws.cell(i, clauseIndex).value.lower():
                    inScreen = False
                    for calc in screenCalcs:
                        ws.insert_rows(i + 1)
                        ws.cell(i + 1, typeIndex, "assign")
                        ws.cell(i + 1, nameIndex, "log") 
                        ws.cell(i + 1, calculationIndex, calc)
                        offset += 1
                    screenCalcs = []


                #Add to the list of screen calculations
                elif ws.cell(i, typeIndex).value != None and ws.cell(i, nameIndex).value != None:
                        name = ws.cell(i, nameIndex).value
                        calc = "JSON.stringify(((data('log')==null)?[]:JSON.parse(data('log'))).concat([[now().toUTCString(),'{0}',data('{0}')]]))".format(name)
                        screenCalcs.append(calc)


            # Not in a single screen section
            else:
                if clauseExists and ws.cell(i, clauseIndex).value != None and 'begin screen' in ws.cell(i, clauseIndex).value.lower():
                    inScreen = True

                elif ws.cell(i, typeIndex).value != None and ws.cell(i, nameIndex).value != None:
                        ws.insert_rows(i + 1)
                        name = ws.cell(i, nameIndex).value
                        ws.cell(i + 1, typeIndex, "assign")
                        ws.cell(i + 1, nameIndex, "log") 
                        # This line is a js instruction which will append the log for this row's data to the main log variable
                        calc = "JSON.stringify(((data('log')==null)?[]:JSON.parse(data('log'))).concat([[now().toUTCString(),'{0}',data('{0}')]]))".format(name)

                        ws.cell(i + 1, calculationIndex, calc)
                        offset += 1


        # Setup the the seperate log rows - we need to split them up as there is a limit on how much data we can submit to the server in one column
        for i in range(LOG_COLUMN_NUMBER):
            rowIndex = ws.max_row + 1
        
            ws.cell(rowIndex, typeIndex, "assign")
            ws.cell(rowIndex, nameIndex, "log_{}".format(i))
            # JS below will ensure that we don't get empty strings which would freeze the app
            calc = "data('log').substring({0},{1}) == '' ? ' ' : data('log').substring({0},{1})".format(i*CHARS_PER_LOG, (i+1)*CHARS_PER_LOG)
            ws.cell(rowIndex, calculationIndex, calc)

        # Fix issue where some entries don't appear for xlsx converter
        for i in range(1, ws.max_row + 1):
            row = ws[i]
            for j in range(1, len(ws[i]) + 1):
                    # Replace None values as they appear as NaN for xlsx converter
                    if ws.cell(i, j).value == None:
                        ws.cell(i, j, "")


        #Set up model table, ensuring that log variables are defined
        if "model" in self.workbook.sheetnames:
            model = self.workbook["model"]
        else:
            model = self.workbook.create_sheet("model")
        
        nameIndex = -1
        typeIndex = -1
        isSessionVariableIndex = -1
        elementTypeIndex = -1

        for i in range(1, model.max_column + 1):
            if model.cell(1, i).value == "name":
                nameIndex = i

            elif model.cell(1, i).value == "type":
                typeIndex = i

            elif model.cell(1, i).value == "isSessionVariable":
                isSessionVariableIndex = i
        
            elif model.cell(1, i).value == "elementType":
                elementTypeIndex = i

        if nameIndex < 0:
            nameIndex = model.max_column + 1
            model.cell(1, nameIndex, "name")

        if typeIndex < 0:
            typeIndex = model.max_column + 1
            model.cell(1, typeIndex, "type")
        
        if isSessionVariableIndex < 0:
            isSessionVariableIndex = model.max_column + 1
            model.cell(1, isSessionVariableIndex, "isSessionVariable")

        if elementTypeIndex < 0:
            elementTypeIndex = model.max_column + 1
            model.cell(1, elementTypeIndex, "elementType")

        #Set up log in model table
        newRow = model.max_row + 1
        model.cell(newRow, nameIndex, "log")
        model.cell(newRow, typeIndex, "string")
        model.cell(newRow, isSessionVariableIndex, "yes")
        model.cell(newRow, elementTypeIndex, "")

        for i in range(LOG_COLUMN_NUMBER):
            newRow = model.max_row + 1
            model.cell(newRow, nameIndex, "log_{}".format(i))
            model.cell(newRow, typeIndex, "string")
            model.cell(newRow, isSessionVariableIndex, "")
            model.cell(newRow, elementTypeIndex, "string({})".format(CHARS_PER_LOG))
        
        print("Saving as {}".format(self.getFileName()))
        self.workbook.save(self.getFileName())


    def getTableId(self):
        try:
            ws = self.workbook['settings'] #Catch error when settings does not exist
        except:
            raise ValueError("Workbook {} must have a sheet named 'settings'".format(self.filename))

        settingIndex = -1
        valueIndex = -1
        for i in range(1, ws.max_column + 1):
            if ws.cell(1, i).value == "setting_name":
                settingIndex = i
            elif ws.cell(1, i).value == "value":
                valueIndex = i

        if settingIndex < 0:
            raise ValueError("Settings sheet in workbook {} must have a column named 'setting_name'".format(self.filename))
        if valueIndex < 0:
            raise ValueError("Settings sheet in workbook {} must have a column named 'value'".format(self.filename))

        tableId = None
        for i in range(1, ws.max_row + 1):
            if ws.cell(i, settingIndex).value == "table_id":
                tableId = ws.cell(i, valueIndex).value
        if tableId == None:
            raise ValueError("Settings sheet in workbook {} must define a value for table_id".format(self.filename))

        return tableId
    

    def getFormId(self):
        try:
            ws = self.workbook['settings'] #Catch error when settings does not exist
        except:
            raise ValueError("Workbook {} must have a sheet named 'settings'".format(self.filename))

        settingIndex = -1
        valueIndex = -1
        for i in range(1, ws.max_column + 1):
            if ws.cell(1, i).value == "setting_name":
                settingIndex = i
            elif ws.cell(1, i).value == "value":
                valueIndex = i

        if settingIndex < 0:
            raise ValueError("Settings sheet in workbook {} must have a column named 'setting_name'".format(self.filename))
        if valueIndex < 0:
            raise ValueError("Settings sheet in workbook {} must have a column named 'value'".format(self.filename))

        formId = None
        for i in range(1, ws.max_row + 1):
            if ws.cell(i, settingIndex).value == "form_id":
                formId = ws.cell(i, valueIndex).value
        if formId == None:
            raise ValueError("Settings sheet in workbook {} must define a value for form_id".format(self.filename))

        return formId


    def getTitle(self):
        try:
            ws = self.workbook['settings'] #Catch error when settings does not exist
        except:
            raise ValueError("Workbook {} must have a sheet named 'settings'".format(self.filename))

        settingIndex = -1
        displayIndex = -1
        for i in range(1, ws.max_column + 1):
            if ws.cell(1, i).value == "setting_name":
                settingIndex = i
            elif ws.cell(1, i).value == "display.title.text.en":
                displayIndex = i

        if settingIndex < 0:
            raise ValueError("Settings sheet in workbook {} must have a column named 'setting_name'".format(self.filename))
        if displayIndex < 0:
            raise ValueError("Settings sheet in workbook {} must have a column named 'display.title.text.en'".format(self.filename))

        title = None
        for i in range(1, ws.max_row + 1):
            if ws.cell(i, settingIndex).value == "survey":
                title = ws.cell(i, displayIndex).value
        if title == None:
            raise ValueError("Settings sheet in workbook {} must define column display.title.text.en for row survey".format(self.filename))

        return title


    def getFileName(self):
        fileName = "designerFiles\\app\\config\\tables\\{0}\\forms\\{1}\\{1}.xlsx".format(self.tableId, self.formId)
        return fileName

    def getJsRepresentation(self):
        js = '{{"formId":"{}", "tableId":"{}", "label":"{}"}}'.format(self.formId, self.tableId, self.title)
        return js

        
def getWorkbooks():
    subDirs = os.listdir("forms")
    workbooks = []
    for directory in subDirs:
        for f in os.listdir("forms\\{}".format(directory)):
            if f[-5:] == ".xlsx":
                workbooks.append("forms\\{}\\{}".format(directory, f))
    return workbooks


def writeJs(content):
    js = "var forms = [{}]".format(','.join(content))
    with open("designerFiles\\app\\config\\tables\\person\\js\\forms.json", "w") as jsFile:
        jsFile.write(js)


if __name__ == "__main__":
    # This javascript describes the forms we want to show in an individual's menu
    jsContent = []
    for workbook in getWorkbooks():
        try:
            print("Converting {}".format(workbook))
            form = Form(workbook)
            form.convert()
            if form.tableId in PERSON_TABLES:
                jsContent.append(form.getJsRepresentation())
            print("Done")
        except Exception as e:
            print("Converting failed due to: {}".format(e))

    writeJs(jsContent)

    os.system("grunt --gruntfile designerFiles\\Gruntfile.js xlsx-convert-all")
